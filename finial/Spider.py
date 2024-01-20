import re
import time
from openpyxl import Workbook
import bs4
import requests


# 模拟浏览器请求头
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/68.0.3440.106 Safari/537.36',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/67.0.3396.99 Safari/537.36',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; …) Gecko/20100101 Firefox/61.0',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/64.0.3282.186 Safari/537.36',
    'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/62.0.3202.62 Safari/537.36',
    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/45.0.2454.101 Safari/537.36',
    'User-Agent': 'Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 6.0)',
    'User-Agent': 'Mozilla/5.0 (Macintosh; U; PPC Mac OS X 10.5; en-US; rv:1.9.2.15) Gecko/20110303 Firefox/3.6.15',
}

# 写入excel
wb = Workbook()
sheet = wb.active
sheet.cell(row=1, column=1).value = "歌曲名"
sheet.cell(row=1, column=2).value = "歌手"
sheet.cell(row=1, column=3).value = "时长"
sheet.cell(row=1, column=4).value = "歌词"

row = 2
# 爬取歌曲名，歌手，时长，歌词
# 做可视化图

txt = ""
singer_count = {}
for page in range(1, 24):
    resp = requests.get(f"https://www.kugou.com/yy/rank/home/1-8888.html?from=rank&p={page}", headers=headers)
    html = bs4.BeautifulSoup(resp.text, "html.parser")
    for li in html.find("div", class_="pc_temp_songlist").find_all("li"):
        try:
            singer, song = li.get("title").strip().split("-")
            singer, song = singer.strip(), song.strip()
            duration = li.find("span", class_="pc_temp_tips_r").find("span",  class_="pc_temp_time").text.strip()
            params = {
                "r": "play/getdata",
                "callback": f"jQuery19105242818611056903_{int(time.time()*1000)}",
                "dfid": "2UHniy2ZkGKQ1ezDT54bwZoZ",
                "appid": "1014",
                "mid": "b4ec8047e1cbcfc610d21610fa9c5abc",
                "platid": "4",
                "encode_album_audio_id": li.get("data-eid"),
                "_": f"{int(time.time()*1000)}"
            }
            resp = requests.get("https://wwwapi.kugou.com/yy/index.php", params=params, headers=headers)
            resp.encoding = "unicode-escape"
            lyrics = re.search('"lyrics":"(?P<lyrics>.*?)",', resp.text, re.S).group("lyrics").replace("\r", "")
            lyrics = " ".join(re.findall("\](?P<lyrics>.*?)\n", lyrics, re.S)).strip()
            txt += lyrics+" "
        except:
            continue
        print({
            "index": row-1,
            "song": song,
            "singer": singer
        })
        sheet.cell(row=row, column=1).value = song
        sheet.cell(row=row, column=2).value = singer
        sheet.cell(row=row, column=3).value = duration
        sheet.cell(row=row, column=4).value = lyrics
        row += 1
    time.sleep(0.15)



wb.save("data.xlsx")
